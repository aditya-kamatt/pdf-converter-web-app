"""
excel_writer.py

Adds a 'product_sizesheet' layout that collapses multiple size rows of the
same product into a single row, with quantities across size columns.

What you get with layout="product_sizesheet"
--------------------------------------------
PRODUCT | 0-3m | 3-6m | 6-12m | 12-18m | 18-24m | 2T | 3T | 4T | 5 | 6 | 7 | 8 | 10 | One size | 0-2T | 2T-4T | 2T-5 | 0-6m | 12-24m | Total

Key behavior
------------
- 'PRODUCT' is derived from Description by stripping the trailing size token.
- All rows with the same PRODUCT are grouped and summed across sizes.
- Size labels are inferred from Description (preferred) or Item SKU suffix.
- Original first-seen order of products is preserved.
- Safe with duplicate input headers; safe regex group names.

Other layouts retained
----------------------
- "standard"     -> a flat 'Orders' sheet (unchanged)
- "sizesheet"    -> one row per style (Item SKU sans size) with sizes across columns (previous behavior)
- "product_sizesheet" -> one row per PRODUCT name (this feature)

Usage
-----
write_to_excel(df, meta, "out.xlsx", layout="product_sizesheet")
"""

from __future__ import annotations

import logging
import re
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# Configuration
# ==============================================================================

# Canonical size order (edit to match your program’s exact buckets)
SIZE_COLUMNS: List[str] = [
    "One size",
    # Adult sizes
    "SS", "S", "M", "L", "XL", "XXL",   
    
    "0-3m", "3-6m", "6-12m", "12-18m", "18-24m",
    "2T", "3T", "4T", "5", "6", "7", "8", "10",
   
    # pack/combined ranges (if present in your data)
    "0-2T", "2T-4T", "2T-5", "0-6m", "12-24m",
]

# Final column titles for descriptive columns per layout
BASE_COLS_MAP: Dict[str, str] = {
    "po_no": "PO #",
    "dev_no": "Dev Code", 
    "item_style": "Item SKU",              # Item SKU with size suffix removed
    "ns_description": "Description",
    "product": "Product",              # Description with trailing size removed
    "hts_code": "HTS Code",
}

# Case-insensitive aliases for incoming headers
COLUMN_ALIASES: Dict[str, Tuple[str, ...]] = {
    "po_no": ("po number", "po #", "po#", "po", "po_no", "po num"),
    "dev_no": ("dev no", "dev #", "dev#", "dev code", "dev_code", "dev", "devcode"),
    "item": ("item", "item sku", "item_sku", "sku"),
    "hts_code": ("hts", "hts code", "hts_code", "hs", "hs code", "hs_code", "htscode"),
    "ns_description": ("ns description", "description", "desc", "item description"),
    "qty": ("qty", "quantity", "order qty", "ordered qty"),
    "size_label": ("size_label", "size", "size name"),  # optional
}

# Tolerant patterns for detecting size tokens in Description
_CANONICAL_TO_PAT = {
    # Adult sizes (tolerant to variants)
    "SS":       r"\b(?:ss|xs|x-?s|extra\s*small|xtra\s*small|ex\s*small)\b",
    "S":        r"\b(?:s|sm|small)\b",
    "M":        r"\b(?:m|med|medium)\b",
    "L":        r"\b(?:l|lg|large)\b",
    "XL":       r"\b(?:xl|x-?l|extra\s*large|xtra\s*large)\b",
    "XXL":      r"\b(?:xxl|2xl|xx-?l|double\s*xl|extra\s*extra\s*large)\b",

    "One size": r"(?:one\s*size|\bos\b)",
    "0-3m":     r"0\s*-\s*3\s*m",
    "3-6m":     r"3\s*-\s*6\s*m",
    "6-12m":    r"6\s*-\s*12\s*m",
    "12-18m":   r"12\s*-\s*18\s*m",
    "18-24m":   r"18\s*-\s*24\s*m",
    "2T":       r"\b2\s*t\b",
    "3T":       r"\b3\s*t\b",
    "4T":       r"\b4\s*t\b",
    "5":        r"\b5\b",
    "6":        r"\b6\b",
    "7":        r"\b7\b",
    "8":        r"\b8\b",
    "10":       r"\b10\b",
    
    # ranges sometimes seen in headers/desc
    "0-2T":     r"0\s*-\s*2\s*t",
    "2T-4T":    r"2\s*t\s*-\s*4\s*t",
    "2T-5":     r"2\s*t\s*-\s*5\b",
    "0-6m":     r"0\s*-\s*6\s*m",
    "12-24m":   r"12\s*-\s*24\s*m",
}

def _build_description_regex():
    """
    Build a safe regex that matches ANY size token in free text.
    Group names are prefixed so they don't start with digits.
    """
    parts, group_to_canon = [], {}
    for canonical, pat in _CANONICAL_TO_PAT.items():
        g = "SIZE_" + re.sub(r"[^A-Za-z0-9]", "_", canonical)  # e.g. "6-12m" -> "SIZE_6_12m"
        parts.append(f"(?P<{g}>{pat})")
        group_to_canon[g] = canonical
    return re.compile("|".join(parts), flags=re.IGNORECASE), group_to_canon

_DESCRIPTION_SIZE_RE, _DESC_GROUP_TO_CANON = _build_description_regex()

# A simple “strip trailing size” regex (handles “- 6-12m”, “(3-6m)”, spaces, dashes)
_TRAILING_SIZE_RE = re.compile(
    r"""
    [\s\-–—]*              # optional space/dash separator
    (?:\(|\b)?             # optional opening parenthesis or word boundary
    (?:
        one\s*size|0\s*-\s*3\s*m|3\s*-\s*6\s*m|6\s*-\s*12\s*m|12\s*-\s*18\s*m|18\s*-\s*24\s*m|
        0\s*-\s*2\s*t|2\s*t\s*-\s*4\s*t|2\s*t\s*-\s*5|0\s*-\s*6\s*m|12\s*-\s*24\s*m|
        \b2\s*t\b|\b3\s*t\b|\b4\s*t\b|\b5\b|\b6\b|\b7\b|\b8\b|\b10\b
    )
    (?:\))?                 # optional closing parenthesis
    \s*$                    # end of string
    """,
    flags=re.IGNORECASE | re.VERBOSE,
)

# SKU suffix → size (fallback if Description lacks size)
SKU_SUFFIX_TO_SIZE: Dict[str, str] = {
    "03M00": "0-3m",
    "0306M": "3-6m",
    "0612M": "6-12m",
    "1218M": "12-18m",
    "1824M": "18-24m",
    "2T000": "2T",
    "3T000": "3T",
    "4T000": "4T",
    "5K000": "5",
    "6K000": "6",
    "7K000": "7",
    "8K000": "8",
    "K0000": "10",
}

# If you still want a style-level sheet elsewhere, this strips known size codes from an Item SKU
_ITEM_SIZE_SUFFIX_RE = re.compile(
    r"(-(?:03M00|0306M|0612M|1218M|1824M|2T000|3T000|4T000|5K000|6K000|7K000|8K000|K0000))$",
    re.IGNORECASE,
)


# ==============================================================================
# Public API
# ==============================================================================

def write_to_excel(
    df: pd.DataFrame,
    meta: dict,
    output_path: str,
    *,
    layout: str = "all",
) -> None:
    """
    Write an Excel workbook with: 
      - 'Summary' sheet
      - 'Orders' sheet (flat)
      - 'SizeSheet' sheet (style-based matrix)
    """
    try:
        # Normalize duplicate headers immediately
        df = _dedupe_columns(df)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            _create_summary_sheet(writer, meta)

            #Orders Sheet
            df.to_excel(writer, sheet_name="Orders", index=False)
            _format_orders_sheet(writer.sheets["Orders"], df)

            #SizeSheet
            prod_df = _to_sizesheet_by_product(df)
            prod_df.to_excel(writer, sheet_name="SizeSheet", index=False)
            _format_product_sizesheet(writer.sheets["SizeSheet"], prod_df)

        logging.info("Excel file written successfully: %s", output_path)
    except Exception as e:
        logging.error("Error writing Excel file: %s", e)
        raise


# ==============================================================================
# Summary sheet
# ==============================================================================

def _create_summary_sheet(writer, meta: dict) -> None:
    summary_data = {
        "Field": [
            "PO Number",
            "Vendor Number",
            "Ship By Date",
            "Payment Terms",
            "Total Amount",
            "Page Count",
            "Processing Date",
        ],
        "Value": [
            meta.get("po_number", "N/A"),
            meta.get("vendor_number", "N/A"),
            meta.get("ship_by_date", "N/A"),
            meta.get("payment_terms", "N/A"),
            f"${meta.get('total', 'N/A')}" if meta.get("total") not in (None, "N/A") else "N/A",
            str(meta.get("page_count", "N/A")),
            pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        ],
    }
    pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)
    _format_summary_sheet(writer.sheets["Summary"])


def _format_summary_sheet(ws) -> None:
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    body_font = Font(size=11)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin


# ==============================================================================
# Orders (flat) — unchanged styling, robust autosize
# ==============================================================================

def _format_orders_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    body_font = Font(size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        s = _series_from_label(df, col) if col in df.columns else None
        if s is not None and len(df) > 0:
            max_len = max(len(str(col)), s.astype(str).str.len().max())
        else:
            max_len = len(str(col))
        ws.column_dimensions[letter].width = min(max_len + 3, 50)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin


# ==============================================================================
# SizeSheet by PRODUCT NAME (one row per product)
# ==============================================================================

def _to_sizesheet_by_product(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a matrix where each PRODUCT (Description without trailing size) is a single row,
    and quantities are spread across size columns.
    """
    if df is None or df.empty:
        cols = [BASE_COLS_MAP["product"]] + SIZE_COLUMNS + ["Total"]
        return pd.DataFrame(columns=cols)

    po_col, dev_col, item_col, desc_col, qty_col, size_col = _map_alias_columns(df)
    hts_col = _pick_hts_column(df)
    required = {"Description": desc_col, "Qty": qty_col}
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError(f"ProductSizeSheet: required columns not found: {', '.join(missing)}")

    work = df.copy()

    # Preserve first-seen product order across the whole file
    work["__row__"] = range(len(work))

    # Derive PRODUCT = Description with the trailing size removed
    desc_series = _series_from_label(work, desc_col).astype(str)
    work["__product__"] = desc_series.map(_strip_trailing_size_from_description)

    # Ensure a normalized size label for pivoting
    if size_col is None or work[size_col].isna().all():
        work["__size__"] = work.apply(
            lambda r: _infer_size(str(r.get(desc_col, "")), str(r.get(item_col, "")) if item_col else ""),
            axis=1,
        )
        size_col = "__size__"
    else:
        size_series = _series_from_label(work, size_col)
        work[size_col] = (
            size_series.astype(str)
            .str.replace("–", "-", regex=False)
            .str.replace("—", "-", regex=False)
            .str.strip()
        )

    # Coerce quantity
    qty_series = _series_from_label(work, qty_col)
    work["__qty__"] = pd.to_numeric(qty_series, errors="coerce").fillna(0).astype(int)

    # First-seen row index per PRODUCT for stable order
    first_seen = work.groupby("__product__", sort=False)["__row__"].min().rename("__first__")
    work = work.merge(first_seen, on="__product__", how="left")

    # Representative (first-seen non-empty) values per PRODUCT
    work["__item_style__"] = _series_from_label(work, item_col).astype(str).apply(_strip_item_size_suffix) if item_col else ""
    rep_cols = {
        "Item SKU": work.groupby("__product__", sort=False)["__item_style__"].apply(_first_nonempty),
        "Dev Code": work.groupby("__product__", sort=False)[dev_col].apply(_first_nonempty) if dev_col else pd.Series(dtype=str),
        "HTS Code": work.groupby("__product__", sort=False)[hts_col].apply(_first_nonempty) if hts_col else pd.Series(dtype=str),
    }
    rep_df = pd.DataFrame(rep_cols).reset_index().rename(columns={"__product__": "__product__"})

    # Pivot: PRODUCT as index, sizes as columns, quantities summed
    pvt = (
        work.pivot_table(
            index="__product__",
            columns=size_col,
            values="__qty__",
            aggfunc="sum",
            fill_value=0,
            observed=False,
        )
        .reset_index()
    )

    # Order rows by first-seen appearance
    pvt = pvt.merge(rep_df, on="__product__", how="left")

    present_sizes = [c for c in pvt.columns if c not in ("__product__", "Item SKU", "Dev Code", "HTS Code")]
    extras = [s for s in present_sizes if s not in SIZE_COLUMNS]
    ordered_sizes = SIZE_COLUMNS + [s for s in extras if s not in SIZE_COLUMNS]

    pvt = (
        pvt.set_index("__product__")
           .reindex(columns=["Item SKU", "Dev Code", "HTS Code"] + ordered_sizes, fill_value=0)
           .reset_index()
    )
    pvt["Total"] = pvt[ordered_sizes].sum(axis=1)

    pvt = pvt.rename(columns={"__product__": BASE_COLS_MAP["product"]})
    pvt = pvt[["Item SKU", "Dev Code", "HTS Code", BASE_COLS_MAP["product"]] + ordered_sizes + ["Total"]]

    return pvt


def _format_product_sizesheet(ws, df: pd.DataFrame) -> None:
    """Styling for ProductSizeSheet: bold header, left-aligned PRODUCT, centered sizes."""
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            if idx in (1, 2, 3, 4):  # PRODUCT
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)

    # Autosize columns robustly
    for col_idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        s = _series_from_label(df, col) if col in df.columns else None
        if s is not None and len(df) > 0:
            max_len = max(len(str(col)), s.astype(str).str.len().max())
        else:
            max_len = len(str(col))
        if col in ("PRODUCT",):
            width = min(max_len + 6, 70)
        elif col in ("Item SKU", "Dev Code", "HTS Code"):
            width = min(max_len + 4, 40)
        elif col == "Total":
            width = max(10, max_len + 2)
        else:
            width = max(6, min(max_len + 1, 10))  # compact size columns
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A2"
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
    zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if r_idx % 2 == 0:
            for cell in row:
                cell.fill = zebra


# ==============================================================================
# Existing style-based SizeSheet (kept for compatibility)
# ==============================================================================

def _to_sizesheet_by_style(df: pd.DataFrame) -> pd.DataFrame:
    """
    Previous behavior: one row per style (Item SKU sans size), sizes across columns.
    """
    po_col, dev_col, item_col, desc_col, qty_col, size_col = _map_alias_columns(df)
    if any(x is None for x in (dev_col, item_col, desc_col, qty_col)):
        raise ValueError("SizeSheet: required columns not found (need Dev Code, Item SKU, Description, Qty).")

    work = df.copy()
    work["__order__"] = range(len(work))
    work["__item_style__"] = work[item_col].astype(str).apply(_strip_item_size_suffix)

    if size_col is None or work[size_col].isna().all():
        work["__size__"] = work.apply(
            lambda r: _infer_size(str(r.get(desc_col, "")), str(r.get(item_col, ""))),
            axis=1,
        )
        size_col = "__size__"
    else:
        size_series = _series_from_label(work, size_col)
        work[size_col] = (
            size_series.astype(str)
            .str.replace("–", "-", regex=False)
            .str.replace("—", "-", regex=False)
            .str.strip()
        )

    qty_series = _series_from_label(work, qty_col)
    work["__qty__"] = pd.to_numeric(qty_series, errors="coerce").fillna(0).astype(int)

    # First occurrence order by style
    first_seen = work.groupby(["__item_style__"], sort=False)["__order__"].min().rename("__first__")
    work = work.merge(first_seen, on="__item_style__", how="left")

    pvt = (
        work.pivot_table(
            index="__item_style__",
            columns=size_col,
            values="__qty__",
            aggfunc="sum",
            fill_value=0,
            observed=False,
        )
        .reset_index()
        .merge(first_seen.reset_index(), on="__item_style__", how="left")
        .sort_values("__first__")
        .drop(columns="__first__")
        .set_index("__item_style__")
        .reindex(columns=SIZE_COLUMNS, fill_value=0)
        .reset_index()
    )

    pvt["Total"] = pvt[SIZE_COLUMNS].sum(axis=1)
    rep_style = work.groupby("__item_style__", sort=False).agg({
        dev_col: _first_nonempty if dev_col else (lambda s: ""),
        (hts_col if (hts_col := _pick_hts_column(work)) else "__dummy__"):_first_nonempty
    }).reset_index()

    #Clean up columns after groupby
    if dev_col:
        rep_style = rep_style.rename(columns={dev_col: "Dev Code"})
    else:
        rep_style["Dev Code"] = ""
    if hts_col:
        rep_style = rep_style.rename(columns={hts_col: "HTS Code"})
    else:
        rep_style["HTS Code"] = ""

    rep_style = rep_style[["__item_style__", "Dev Code", "HTS Code"]]
    pvt = pvt.merge(rep_style, on="__item_style__", how="left")
    pvt = pvt.rename(columns={"__item_style__": BASE_COLS_MAP["item_style"]})
    leading = [BASE_COLS_MAP["item_style"], "Dev Code", "HTS Code"]
    pvt = pvt[leading + SIZE_COLUMNS + ["Total"]]
    return pvt


def _format_sizesheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            if idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)
    for col_idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        s = _series_from_label(df, col) if col in df.columns else None
        if s is not None and len(df) > 0:
            max_len = max(len(str(col)), s.astype(str).str.len().max())
        else:
            max_len = len(str(col))
        if col in (BASE_COLS_MAP["item_style"], "Dev Code","HTS Code"):
            width = min(max_len + 6, 60)
        elif col == "Total":
            width = max(10, max_len + 2)
        else:
            width = max(6, min(max_len + 1, 10))
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
    zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if r_idx % 2 == 0:
            for cell in row:
                cell.fill = zebra


# ==============================================================================
# Helpers (robustness, aliasing, parsing)
# ==============================================================================

def _dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Make incoming headers unique by appending .1, .2, ... to duplicates (first kept as-is)."""
    out = df.copy()
    seen, new_cols = {}, []
    for c in out.columns:
        k = str(c)
        if k in seen:
            seen[k] += 1
            new_cols.append(f"{k}.{seen[k]}")
        else:
            seen[k] = 0
            new_cols.append(k)
    out.columns = new_cols
    return out


def _series_from_label(df: pd.DataFrame, label) -> Optional[pd.Series]:
    """
    Return a single Series for a label. If the label is duplicated (df[label] is a DataFrame),
    take the first column.
    """
    obj = df[label]
    if isinstance(obj, pd.DataFrame):
        return obj.iloc[:, 0]
    return obj


def _map_alias_columns(df: pd.DataFrame):
    """
    Return actual column names for (po_no, dev_no, item, ns_description, qty, size_label)
    using case-insensitive aliases. First occurrence wins when duplicated.
    """
    lookup = {}
    for c in df.columns:
        k = str(c).lower().strip()
        if k not in lookup:
            lookup[k] = c

    def pick(opts: Iterable[str]) -> Optional[str]:
        for o in opts:
            if o in lookup:
                return lookup[o]
        return None

    po_col   = pick(COLUMN_ALIASES.get("po_no", ()))
    dev_col  = pick(COLUMN_ALIASES["dev_no"])
    item_col = pick(COLUMN_ALIASES["item"])
    desc_col = pick(COLUMN_ALIASES["ns_description"])
    qty_col  = pick(COLUMN_ALIASES["qty"])
    size_col = pick(COLUMN_ALIASES.get("size_label", ()))  # optional

    return po_col, dev_col, item_col, desc_col, qty_col, size_col

def _pick_hts_column(df: pd.DataFrame) -> Optional[str]:
    lookup = {str(c).lower().strip(): c for c in df.columns}
    for alias in COLUMN_ALIASES.get("hts_code", ()):
        if alias in lookup:
            return lookup[alias]
    return None

def _infer_size_from_text(text: str) -> str:
    """Pick the first matching size token from free text using the description regex."""
    m = _DESCRIPTION_SIZE_RE.search(text or "")
    if not m:
        return "Unknown"
    for gname, val in m.groupdict().items():
        if val:
            return _DESC_GROUP_TO_CANON.get(gname, "Unknown")
    return "Unknown"


def _infer_size(description: str, item_sku: str) -> str:
    """Infer size from Description (preferred) or Item SKU suffix (fallback)."""
    desc_norm = (description or "").lower().replace("–", "-").replace("—", "-")
    size = _infer_size_from_text(desc_norm)
    if size != "Unknown":
        return size
    sku = (item_sku or "").strip().upper()
    suf5, suf6 = sku[-5:], sku[-6:]
    if suf5 in SKU_SUFFIX_TO_SIZE:
        return SKU_SUFFIX_TO_SIZE[suf5]
    for k, v in SKU_SUFFIX_TO_SIZE.items():
        if suf6.endswith(k):
            return v
    return "Unknown"

def _strip_item_size_suffix(item_sku: str) -> str:
    """Remove a known trailing size code from an Item SKU to get the STYLE code."""
    if not isinstance(item_sku, str):
        return str(item_sku)
    return _ITEM_SIZE_SUFFIX_RE.sub("", item_sku).strip()


def _strip_trailing_size_from_description(desc: str) -> str:
    """
    Remove a trailing size token from the description to derive the PRODUCT name.
    Examples:
      "Pink Ruffle Trim Woven Shorts - 6-12m" -> "Pink Ruffle Trim Woven Shorts"
      "Boys Swim Trunks (3-6m)" -> "Boys Swim Trunks"
    """
    if not isinstance(desc, str):
        return str(desc)
    d = desc.strip().replace("–", "-").replace("—", "-")
    d = _TRAILING_SIZE_RE.sub("", d)
    return d.strip(" -–—").strip()
def _first_nonempty(s: pd.Series) -> str:
    for v in s:
        if pd.notna(v) and str(v).strip() != "":
            return str(v)
    return ""
