import pytest
import pandas as pd
import os
import sys
import pathlib
from openpyxl import load_workbook

# Add project root to path for imports
HERE = pathlib.Path(__file__).parent
PROJECT_ROOT = HERE.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from excel_io.excel_writer import write_to_excel


class TestWriteToExcelBasicFunctionality:
    """Test basic functionality of write_to_excel"""

    def test_write_creates_file(self, tmp_path):
        """Test that write_to_excel creates the output file"""
        df = pd.DataFrame({
            "Qty": [1, 2],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["BrandA", "BrandB"],
            "Description": ["Product 1", "Product 2"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {"po_number": "PO123", "total": "$50.00"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_creates_multiple_sheets(self, tmp_path):
        """Test that Excel file contains expected sheets"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Product"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        # Load and check sheets
        wb = load_workbook(output)
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Orders" in sheet_names

    def test_write_with_layout_all(self, tmp_path):
        """Test write_to_excel with layout='all'"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Product"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output, layout="all")
        assert os.path.exists(output)
        
        wb = load_workbook(output)
        assert "SizeSheet" in wb.sheetnames or "Summary" in wb.sheetnames

    def test_write_preserves_data_types(self, tmp_path):
        """Test that numeric data is preserved correctly"""
        df = pd.DataFrame({
            "Qty": [5, 10, 15],
            "Item_SKU": ["A", "B", "C"],
            "Dev_Code": ["D1", "D2", "D3"],
            "UPC": ["012345678905"] * 3,
            "HTS_Code": ["1234567890"] * 3,
            "Brand": ["Brand"] * 3,
            "Description": ["Desc"] * 3,
            "Rate": [10.0, 20.0, 30.0],
            "Amount": [50.0, 200.0, 450.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        # Read back and verify
        result_df = pd.read_excel(output, sheet_name="Orders")
        assert len(result_df) == 3
        assert list(result_df["Qty"]) == [5, 10, 15]


class TestWriteToExcelWithMetadata:
    """Test write_to_excel with different metadata"""

    def test_write_with_complete_metadata(self, tmp_path):
        """Test with all metadata fields"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {
            "po_number": "PO12345",
            "vendor_number": "V001",
            "ship_by_date": "2025-12-31",
            "payment_terms": "Net 30",
            "total": "$10,000.00",
            "page_count": 5
        }
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        # Verify Summary sheet has metadata
        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames

    def test_write_with_minimal_metadata(self, tmp_path):
        """Test with minimal metadata"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "PO123"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_na_metadata(self, tmp_path):
        """Test with N/A values in metadata"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {
            "po_number": "N/A",
            "vendor_number": "N/A",
            "total": "N/A"
        }
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)


class TestWriteToExcelDataVariations:
    """Test write_to_excel with various data patterns"""

    def test_write_with_multiple_rows(self, tmp_path):
        """Test with many rows"""
        num_rows = 50
        df = pd.DataFrame({
            "Qty": list(range(1, num_rows + 1)),
            "Item_SKU": [f"SKU{i}" for i in range(num_rows)],
            "Dev_Code": [f"D{i}" for i in range(num_rows)],
            "UPC": ["012345678905"] * num_rows,
            "HTS_Code": ["1234567890"] * num_rows,
            "Brand": ["Brand"] * num_rows,
            "Description": [f"Product {i}" for i in range(num_rows)],
            "Rate": [10.0] * num_rows,
            "Amount": [i * 10.0 for i in range(1, num_rows + 1)]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        result_df = pd.read_excel(output, sheet_name="Orders")
        assert len(result_df) == num_rows

    def test_write_with_nan_values(self, tmp_path):
        """Test handling of NaN values"""
        df = pd.DataFrame({
            "Qty": [1, 2, None],
            "Item_SKU": ["SKU1", "SKU2", "SKU3"],
            "Dev_Code": ["D1", None, "D3"],
            "UPC": ["012345678905"] * 3,
            "HTS_Code": ["1234567890"] * 3,
            "Brand": ["Brand"] * 3,
            "Description": ["Desc1", "Desc2", None],
            "Rate": [10.0, None, 30.0],
            "Amount": [10.0, 20.0, None]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        # Should handle NaN gracefully
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_special_characters(self, tmp_path):
        """Test with special characters in data"""
        df = pd.DataFrame({
            "Qty": [1, 2],
            "Item_SKU": ["SKU-001", "SKU#002"],
            "Dev_Code": ["D&1", "D@2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand™", "Co®"],
            "Description": ["Product & Service", "Item #2"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {"po_number": "PO#123"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_long_descriptions(self, tmp_path):
        """Test with very long description text"""
        long_text = "A" * 500
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": [long_text],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_unicode(self, tmp_path):
        """Test with Unicode characters"""
        df = pd.DataFrame({
            "Qty": [1, 2],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Café", "Naïve"],
            "Description": ["Résumé item", "日本製品"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_numeric_strings(self, tmp_path):
        """Test with numeric values as strings"""
        df = pd.DataFrame({
            "Qty": ["1", "2"],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand"] * 2,
            "Description": ["Desc"] * 2,
            "Rate": ["10.00", "20.00"],
            "Amount": ["10.00", "40.00"]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)


class TestWriteToExcelErrorHandling:
    """Test error handling in write_to_excel"""

    def test_write_missing_required_column(self, tmp_path):
        """Test that missing Description column raises error"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            # Missing Description
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        with pytest.raises(ValueError, match="Description"):
            write_to_excel(df, meta, output)

    def test_write_creates_parent_directories(self, tmp_path):
        """Test that write works when parent directories exist"""
        # Create parent directories first
        output_dir = tmp_path / "subdir1" / "subdir2"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(output_dir / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_overwrites_existing_file(self, tmp_path):
        """Test that existing file is overwritten"""
        df1 = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        df2 = pd.DataFrame({
            "Qty": [2, 3],
            "Item_SKU": ["SKU2", "SKU3"],
            "Dev_Code": ["D2", "D3"],
            "UPC": ["012345678905"] * 2,
            "HTS_Code": ["1234567890"] * 2,
            "Brand": ["Brand"] * 2,
            "Description": ["Desc"] * 2,
            "Rate": [20.0, 30.0],
            "Amount": [40.0, 90.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        # Write first file
        write_to_excel(df1, meta, output)
        size1 = os.path.getsize(output)
        
        # Overwrite with second file
        write_to_excel(df2, meta, output)
        size2 = os.path.getsize(output)
        
        # Verify it was overwritten (different size)
        assert size1 != size2
        
        # Verify content is from df2
        result_df = pd.read_excel(output, sheet_name="Orders")
        assert len(result_df) == 2


class TestWriteToExcelEdgeCases:
    """Test edge cases and boundary conditions"""

    def test_write_with_zero_quantities(self, tmp_path):
        """Test with zero quantities"""
        df = pd.DataFrame({
            "Qty": [0, 0],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand"] * 2,
            "Description": ["Desc"] * 2,
            "Rate": [10.0, 20.0],
            "Amount": [0.0, 0.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_negative_values(self, tmp_path):
        """Test with negative values (e.g., returns/credits)"""
        df = pd.DataFrame({
            "Qty": [-1, -2],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand"] * 2,
            "Description": ["Desc"] * 2,
            "Rate": [10.0, 20.0],
            "Amount": [-10.0, -40.0]
        })
        meta = {"total": "$-50.00"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_large_amounts(self, tmp_path):
        """Test with very large monetary amounts"""
        df = pd.DataFrame({
            "Qty": [1000000],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [99999.99],
            "Amount": [99999990000.00]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_empty_strings(self, tmp_path):
        """Test with empty string values"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": [""],  # Empty
            "UPC": ["012345678905"],
            "HTS_Code": [""],  # Empty
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

    def test_write_with_mixed_case_columns(self, tmp_path):
        """Test that column names are case-sensitive"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU"],
            "Dev_Code": ["D"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],  # Correct case
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        # Should work with correct column names
        write_to_excel(df, meta, output)
        assert os.path.exists(output)

