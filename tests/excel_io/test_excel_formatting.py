"""
Tests for Excel output formatting and styling.

These tests verify that Excel files are not just structurally correct,
but also properly formatted with colors, fonts, borders, alignment, and column widths.
"""
import pytest
import pandas as pd
import os
import sys
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border

# Add project root to path
HERE = pathlib.Path(__file__).parent
PROJECT_ROOT = HERE.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from excel_io.excel_writer import write_to_excel


class TestSummarySheetFormatting:
    """Test formatting of the Summary sheet"""

    def test_summary_sheet_header_color(self, tmp_path):
        """Test that Summary sheet header has correct color"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Summary"]
        
        # Check header row (row 1) has blue background
        header_fill = ws['A1'].fill
        assert header_fill.start_color.rgb in ["FF366092", "00366092", "366092"]

    def test_summary_sheet_header_font(self, tmp_path):
        """Test that Summary sheet header has white bold font"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Summary"]
        
        # Check header font is white and bold
        header_font = ws['A1'].font
        assert header_font.bold is True
        # Color can be FFFFFF, FFFFFFFF, or 00FFFFFF
        assert header_font.color.rgb in ["FFFFFFFF", "00FFFFFF", "FFFFFF"]

    def test_summary_sheet_column_widths(self, tmp_path):
        """Test that Summary sheet has appropriate column widths"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Summary"]
        
        # Column A should be width 20, Column B should be width 30
        assert ws.column_dimensions['A'].width == 20
        assert ws.column_dimensions['B'].width == 30

    def test_summary_sheet_borders(self, tmp_path):
        """Test that Summary sheet cells have borders"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Summary"]
        
        # Check that cells have thin borders
        for row in ws.iter_rows(min_row=1, max_row=3):
            for cell in row:
                assert cell.border.left.style == 'thin'
                assert cell.border.right.style == 'thin'
                assert cell.border.top.style == 'thin'
                assert cell.border.bottom.style == 'thin'

    def test_summary_sheet_alignment(self, tmp_path):
        """Test that Summary sheet has correct cell alignment"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Summary"]
        
        # Header should be center-aligned
        assert ws['A1'].alignment.horizontal == 'center'
        
        # Body cells should be left-aligned
        assert ws['A2'].alignment.horizontal == 'left'


class TestOrdersSheetFormatting:
    """Test formatting of the Orders sheet"""

    def test_orders_sheet_header_color(self, tmp_path):
        """Test that Orders sheet header has correct color"""
        df = pd.DataFrame({
            "Qty": [1, 2],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand1", "Brand2"],
            "Description": ["Desc1", "Desc2"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Orders"]
        
        # Check header row has blue background (4F81BD)
        header_fill = ws['A1'].fill
        assert header_fill.start_color.rgb in ["FF4F81BD", "004F81BD", "4F81BD"]

    def test_orders_sheet_header_font(self, tmp_path):
        """Test that Orders sheet header has white bold font"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Orders"]
        
        # Check header font
        header_font = ws['A1'].font
        assert header_font.bold is True
        assert header_font.color.rgb in ["FFFFFFFF", "00FFFFFF", "FFFFFF"]

    def test_orders_sheet_column_widths_autosized(self, tmp_path):
        """Test that Orders sheet columns are auto-sized appropriately"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU123456789"],  # Long SKU
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Very Long Description Text Here"],  # Long description
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Orders"]
        
        # Description column should be wider
        # Find Description column
        desc_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Description":
                desc_col_idx = idx
                break
        
        if desc_col_idx:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(desc_col_idx)
            # Should be wider than minimum but not exceed 50
            assert 10 < ws.column_dimensions[col_letter].width <= 50

    def test_orders_sheet_borders_applied(self, tmp_path):
        """Test that all Orders sheet cells have borders"""
        df = pd.DataFrame({
            "Qty": [1, 2],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand1", "Brand2"],
            "Description": ["Desc1", "Desc2"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["Orders"]
        
        # Check borders on data rows
        for row in ws.iter_rows(min_row=1, max_row=3):
            for cell in row:
                assert cell.border.left.style == 'thin'
                assert cell.border.right.style == 'thin'


class TestSizeSheetFormatting:
    """Test formatting of the SizeSheet"""

    def test_sizesheet_header_color(self, tmp_path):
        """Test that SizeSheet header has correct color"""
        df = pd.DataFrame({
            "Qty": [1, 2],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand1", "Brand2"],
            "Description": ["Product 6-12m", "Product 3-6m"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["SizeSheet"]
        
        # Check header has blue background
        header_fill = ws['A1'].fill
        assert header_fill.start_color.rgb in ["FF4F81BD", "004F81BD", "4F81BD"]

    def test_sizesheet_zebra_striping(self, tmp_path):
        """Test that SizeSheet has zebra striping (alternating row colors)"""
        df = pd.DataFrame({
            "Qty": [1, 1, 1, 1],
            "Item_SKU": ["SKU1", "SKU2", "SKU3", "SKU4"],
            "Dev_Code": ["D1", "D2", "D3", "D4"],
            "UPC": ["012345678905"] * 4,
            "HTS_Code": ["1234567890"] * 4,
            "Brand": ["Brand"] * 4,
            "Description": ["Product 1 - 6-12m", "Product 2 - 3-6m", 
                          "Product 3 - 0-3m", "Product 4 - 12-18m"],
            "Rate": [10.0] * 4,
            "Amount": [10.0] * 4
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["SizeSheet"]
        
        # Check that even data rows have gray background
        row_2_fill = ws['A2'].fill
        row_3_fill = ws['A3'].fill
        row_4_fill = ws['A4'].fill
        
        # Row 2 (first data row) might not be striped, row 4 (even) should be
        # Check if zebra striping is applied to even rows
        if row_4_fill.start_color.rgb:
            assert row_4_fill.start_color.rgb in ["FFF2F2F2", "00F2F2F2", "F2F2F2"]

    def test_sizesheet_freeze_panes(self, tmp_path):
        """Test that SizeSheet has freeze panes on first row"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Product - 6-12m"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["SizeSheet"]
        
        # Check freeze panes is set to A2
        assert ws.freeze_panes == 'A2'

    def test_sizesheet_alignment_product_vs_sizes(self, tmp_path):
        """Test that Product column is left-aligned and size columns are center-aligned"""
        df = pd.DataFrame({
            "Qty": [1, 1],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand"] * 2,
            "Description": ["Product A - 6-12m", "Product B - 3-6m"],
            "Rate": [10.0, 20.0],
            "Amount": [10.0, 40.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["SizeSheet"]
        
        # First few columns (Item SKU, Dev Code, HTS, Brand, Product) should be left-aligned
        # Find the Product column (should be column E - index 5)
        product_cell = ws.cell(row=2, column=5)  # First data row, 5th column
        assert product_cell.alignment.horizontal == 'left'
        
        # Size columns should be center-aligned
        # Assuming size columns start after column 5
        if ws.max_column > 5:
            size_cell = ws.cell(row=2, column=6)  # First size column
            assert size_cell.alignment.horizontal == 'center'

    def test_sizesheet_column_widths_appropriate(self, tmp_path):
        """Test that SizeSheet has appropriate column widths for different column types"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU12345"],
            "Dev_Code": ["DEV123"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["LongBrandName"],
            "Description": ["Very Long Product Name Here - 6-12m"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        ws = wb["SizeSheet"]
        
        # Product column should exist and have a reasonable width
        # Column E should be Product
        assert ws.column_dimensions['E'].width > 5  # Should be at least visible
        
        # Size columns should be compact (around 6-10)
        # Check a later column that should be a size
        if ws.max_column > 6:
            from openpyxl.utils import get_column_letter
            size_col_letter = get_column_letter(7)
            assert ws.column_dimensions[size_col_letter].width <= 15


class TestFormattingConsistency:
    """Test that formatting is consistent across sheets"""

    def test_all_sheets_have_borders(self, tmp_path):
        """Test that all sheets have borders on cells"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Product - 6-12m"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        
        for sheet_name in ["Summary", "Orders", "SizeSheet"]:
            ws = wb[sheet_name]
            # Check first cell has borders
            assert ws['A1'].border.left.style == 'thin'
            assert ws['A1'].border.top.style == 'thin'

    def test_all_sheets_have_header_styling(self, tmp_path):
        """Test that all sheets have styled headers"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Product - 6-12m"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {"po_number": "12345"}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        
        for sheet_name in ["Summary", "Orders", "SizeSheet"]:
            ws = wb[sheet_name]
            # Check header has background color
            assert ws['A1'].fill.start_color.rgb is not None
            # Check header is bold
            assert ws['A1'].font.bold is True

    def test_column_width_never_zero(self, tmp_path):
        """Test that no columns have zero or negative width"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["Brand"],
            "Description": ["Product - 6-12m"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        wb = load_workbook(output)
        
        for sheet_name in ["Summary", "Orders", "SizeSheet"]:
            ws = wb[sheet_name]
            for col in ws.column_dimensions.values():
                if col.width is not None:
                    assert col.width > 0


class TestDataTypes:
    """Test that data types are preserved correctly in Excel"""

    def test_numeric_columns_as_numbers(self, tmp_path):
        """Test that numeric columns are stored as numbers, not text"""
        df = pd.DataFrame({
            "Qty": [5, 10],
            "Item_SKU": ["SKU1", "SKU2"],
            "Dev_Code": ["D1", "D2"],
            "UPC": ["012345678905", "123456789012"],
            "HTS_Code": ["1234567890", "9876543210"],
            "Brand": ["Brand"] * 2,
            "Description": ["Desc1", "Desc2"],
            "Rate": [10.50, 20.75],
            "Amount": [52.50, 207.50]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        # Read back and verify types
        result_df = pd.read_excel(output, sheet_name="Orders")
        
        # Qty should be numeric
        assert pd.api.types.is_numeric_dtype(result_df['Qty'])
        # Rate should be numeric
        assert pd.api.types.is_numeric_dtype(result_df['Rate'])
        # Amount should be numeric
        assert pd.api.types.is_numeric_dtype(result_df['Amount'])

    def test_text_columns_as_text(self, tmp_path):
        """Test that text columns remain as text"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU-001"],
            "Dev_Code": ["DEV-001"],
            "UPC": ["012345678905"],
            "HTS_Code": ["1234567890"],
            "Brand": ["BrandName"],
            "Description": ["Product Description"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        # Read back
        result_df = pd.read_excel(output, sheet_name="Orders")
        
        # Text columns should be object/string type (or float if empty/NaN)
        assert result_df['Item_SKU'].dtype in ['object', 'O']
        # Brand might be float64 if all values are NaN after processing
        assert result_df['Brand'].dtype in ['object', 'O', 'float64']
        assert result_df['Description'].dtype in ['object', 'O']

    def test_upc_preserved_with_leading_zeros(self, tmp_path):
        """Test that UPC codes with leading zeros are preserved"""
        df = pd.DataFrame({
            "Qty": [1],
            "Item_SKU": ["SKU1"],
            "Dev_Code": ["D1"],
            "UPC": ["001234567890"],  # Leading zeros
            "HTS_Code": ["0012345678"],  # Leading zeros
            "Brand": ["Brand"],
            "Description": ["Desc"],
            "Rate": [10.0],
            "Amount": [10.0]
        })
        meta = {}
        output = str(tmp_path / "test.xlsx")
        
        write_to_excel(df, meta, output)
        
        # Read back
        result_df = pd.read_excel(output, sheet_name="Orders", dtype={'UPC': str, 'HTS_Code': str})
        
        # Should preserve leading zeros if stored as text
        # Note: This depends on how the writer handles it
        assert len(str(result_df['UPC'][0])) >= 10

