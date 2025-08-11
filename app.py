from flask import Flask, request, jsonify, send_file, render_template, flash, redirect, url_for
from werkzeug.utils import secure_filename
import tempfile
import os
import logging
from datetime import datetime
import pandas as pd

# Import your existing modules with corrected paths
from core.parser import extract_header_meta, extract_table_rows
from excel_io.excel_writer import write_to_excel
from pdf_utils.logging_config import setup_logging

# Setup logging
setup_logging()
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'  # Change this in production
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'pdf'}
UPLOAD_FOLDER = 'temp_uploads'
OUTPUT_FOLDER = 'temp_outputs'

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if the uploaded file is a PDF."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_data_from_pdf(pdf_path):
    """
    Extract data from PDF using your existing parser logic.
    Returns meta and DataFrame.
    """
    try:
        # Extract metadata
        meta = extract_header_meta(pdf_path)
        logger.info(f"Extracted metadata: {meta}")
        
        # Extract table rows
        rows = extract_table_rows(pdf_path)
        
        if not rows or len(rows) <= 1:
            raise ValueError("No data rows found in PDF")
        
        # Convert to DataFrame
        header = rows[0]
        data_rows = rows[1:]
        df = pd.DataFrame(data_rows, columns=header)
        
        logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
        return meta, df
        
    except Exception as e:
        logger.error(f"Error extracting data from PDF: {str(e)}")
        raise

@app.route('/')
def index():
    """Main upload page."""
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert_pdf():
    """Handle PDF conversion."""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        if not allowed_file(file.filename):
            flash('Please upload a PDF file', 'error')
            return redirect(url_for('index'))
        
        # Secure the filename
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save uploaded file
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        logger.info(f"Saved uploaded file to: {pdf_path}")
        
        # Extract data from PDF
        meta, df = extract_data_from_pdf(pdf_path)
        
        # Create output Excel file
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        # Write to Excel using your existing function
        write_to_excel(df, meta, excel_path)
        
        logger.info(f"Created Excel file: {excel_path}")
        
        # Clean up uploaded PDF
        try:
            os.remove(pdf_path)
        except OSError:
            logger.warning(f"Could not remove uploaded file: {pdf_path}")
        
        # Return the Excel file
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"converted_{filename.rsplit('.', 1)[0]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error during conversion: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/api/convert', methods=['POST'])
def api_convert():
    """API endpoint for programmatic access."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        # Process file similar to web interface
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        # Extract and convert
        meta, df = extract_data_from_pdf(pdf_path)
        
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        write_to_excel(df, meta, excel_path)
        
        # Clean up
        try:
            os.remove(pdf_path)
        except OSError:
            pass
        
        # Return success response with metadata
        return jsonify({
            'success': True,
            'message': 'File converted successfully',
            'metadata': meta,
            'rows_processed': len(df),
            'download_url': f'/download/{excel_filename}'
        })
        
    except Exception as e:
        logger.error(f"API conversion error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """Download converted files."""
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'error': 'Error downloading file'}), 500

@app.route('/health')
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

@app.errorhandler(413)
def too_large(e):
    flash('File too large. Maximum size is 16MB.', 'error')
    return redirect(url_for('index'))

@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Unhandled exception: {str(e)}")
    flash('An unexpected error occurred. Please try again.', 'error')
    return redirect(url_for('index'))

# Cleanup old files periodically (you might want to run this as a background task)
def cleanup_old_files():
    """Remove files older than 1 hour."""
    import time
    current_time = time.time()
    
    for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            if os.path.isfile(file_path):
                file_age = current_time - os.path.getctime(file_path)
                if file_age > 3600:  # 1 hour
                    try:
                        os.remove(file_path)
                        logger.info(f"Cleaned up old file: {file_path}")
                    except OSError:
                        logger.warning(f"Could not remove old file: {file_path}")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)rom flask import Flask, request, jsonify, send_file, render_template, flash, redirect, url_for
from werkzeug.utils import secure_filename
import os
import logging
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import your existing modules
from core.parser import extract_header_meta, extract_table_rows
from pdf_utils.logging_config import setup_logging

# Setup logging
setup_logging()
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'pdf'}
UPLOAD_FOLDER = 'temp_uploads'
OUTPUT_FOLDER = 'temp_outputs'

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if the uploaded file is a PDF."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def create_excel_from_rows(rows, meta, output_path):
    """Create Excel file without pandas dependency."""
    wb = Workbook()
    
    # Create Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    # Add summary data
    summary_data = [
        ['Field', 'Value'],
        ['PO Number', meta.get('po_number', 'N/A')],
        ['Vendor Number', meta.get('vendor_number', 'N/A')],
        ['Ship By Date', meta.get('ship_by_date', 'N/A')],
        ['Payment Terms', meta.get('payment_terms', 'N/A')],
        ['Total Amount', f"${meta.get('total', 'N/A')}" if meta.get('total') != 'N/A' else 'N/A'],
        ['Page Count', str(meta.get('page_count', 'N/A'))],
        ['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Format summary sheet
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 30
    
    # Create Orders sheet
    orders_ws = wb.create_sheet(title="Orders")
    
    # Add data rows
    for row in rows:
        orders_ws.append(row)
    
    # Format orders sheet header
    if rows:
        for cell in orders_ws[1]:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in orders_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set specific widths based on column content
        if column_letter == 'A':  # Qty
            adjusted_width = max(max_length + 2, 8)
        elif column_letter in ['B', 'C']:  # Item SKU, Dev Code
            adjusted_width = max(max_length + 2, 15)
        elif column_letter in ['D', 'E']:  # UPC, HTS Code
            adjusted_width = max(max_length + 2, 18)
        elif column_letter == 'F':  # Brand
            adjusted_width = max(max_length + 2, 15)
        elif column_letter == 'G':  # Description
            adjusted_width = min(max(max_length + 2, 20), 50)  # Cap at 50
        elif column_letter in ['H', 'I']:  # Rate, Amount
            adjusted_width = max(max_length + 2, 12)
        else:
            adjusted_width = max_length + 2
            
        orders_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)
    logger.info(f"Excel file created: {output_path}")

def extract_data_from_pdf(pdf_path):
    """Extract data from PDF using existing parser."""
    try:
        # Extract metadata
        meta = extract_header_meta(pdf_path)
        logger.info(f"Extracted metadata: {meta}")
        
        # Extract table rows
        rows = extract_table_rows(pdf_path)
        
        if not rows or len(rows) <= 1:
            raise ValueError("No data rows found in PDF")
        
        logger.info(f"Extracted {len(rows)} rows from PDF")
        return meta, rows
        
    except Exception as e:
        logger.error(f"Error extracting data from PDF: {str(e)}")
        raise

@app.route('/')
def index():
    """Main upload page."""
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert_pdf():
    """Handle PDF conversion."""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        if not allowed_file(file.filename):
            flash('Please upload a PDF file', 'error')
            return redirect(url_for('index'))
        
        # Secure the filename
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save uploaded file
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        logger.info(f"Saved uploaded file to: {pdf_path}")
        
        # Extract data from PDF
        meta, rows = extract_data_from_pdf(pdf_path)
        
        # Create output Excel file
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        # Create Excel file
        create_excel_from_rows(rows, meta, excel_path)
        
        logger.info(f"Created Excel file: {excel_path}")
        
        # Clean up uploaded PDF
        try:
            os.remove(pdf_path)
        except OSError:
            logger.warning(f"Could not remove uploaded file: {pdf_path}")
        
        # Return the Excel file
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"converted_{filename.rsplit('.', 1)[0]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error during conversion: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/api/convert', methods=['POST'])
def api_convert():
    """API endpoint for programmatic access."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        # Process file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        # Extract and convert
        meta, rows = extract_data_from_pdf(pdf_path)
        
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        create_excel_from_rows(rows, meta, excel_path)
        
        # Clean up
        try:
            os.remove(pdf_path)
        except OSError:
            pass
        
        # Return success response
        return jsonify({
            'success': True,
            'message': 'File converted successfully',
            'metadata': meta,
            'rows_processed': len(rows) - 1,  # Subtract header row
            'download_url': f'/download/{excel_filename}'
        })
        
    except Exception as e:
        logger.error(f"API conversion error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """Download converted files."""
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'error': 'Error downloading file'}), 500

@app.route('/health')
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)rom flask import Flask, request, jsonify, send_file, render_template, flash, redirect, url_for
from werkzeug.utils import secure_filename
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import your existing modules
from core.parser import extract_header_meta, extract_table_rows
from pdf_utils.logging_config import setup_logging

# Setup logging
setup_logging()
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'pdf'}
UPLOAD_FOLDER = 'temp_uploads'
OUTPUT_FOLDER = 'temp_outputs'

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if the uploaded file is a PDF."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def create_excel_from_rows(rows, meta, output_path):
    """Create Excel file without pandas dependency."""
    wb = Workbook()
    
    # Create Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    # Add summary data
    summary_data = [
        ['Field', 'Value'],
        ['PO Number', meta.get('po_number', 'N/A')],
        ['Vendor Number', meta.get('vendor_number', 'N/A')],
        ['Ship By Date', meta.get('ship_by_date', 'N/A')],
        ['Payment Terms', meta.get('payment_terms', 'N/A')],
        ['Total Amount', f"${meta.get('total', 'N/A')}" if meta.get('total') != 'N/A' else 'N/A'],
        ['Page Count', str(meta.get('page_count', 'N/A'))],
        ['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Format summary sheet
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 30
    
    # Create Orders sheet
    orders_ws = wb.create_sheet(title="Orders")
    
    # Add data rows
    for row in rows:
        orders_ws.append(row)
    
    # Format orders sheet header
    if rows:
        for cell in orders_ws[1]:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in orders_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set specific widths based on column content
        if column_letter == 'A':  # Qty
            adjusted_width = max(max_length + 2, 8)
        elif column_letter in ['B', 'C']:  # Item SKU, Dev Code
            adjusted_width = max(max_length + 2, 15)
        elif column_letter in ['D', 'E']:  # UPC, HTS Code
            adjusted_width = max(max_length + 2, 18)
        elif column_letter == 'F':  # Brand
            adjusted_width = max(max_length + 2, 15)
        elif column_letter == 'G':  # Description
            adjusted_width = min(max(max_length + 2, 20), 50)  # Cap at 50
        elif column_letter in ['H', 'I']:  # Rate, Amount
            adjusted_width = max(max_length + 2, 12)
        else:
            adjusted_width = max_length + 2
            
        orders_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)
    logger.info(f"Excel file created: {output_path}")

def extract_data_from_pdf(pdf_path):
    """Extract data from PDF using existing parser."""
    try:
        # Extract metadata
        meta = extract_header_meta(pdf_path)
        logger.info(f"Extracted metadata: {meta}")
        
        # Extract table rows
        rows = extract_table_rows(pdf_path)
        
        if not rows or len(rows) <= 1:
            raise ValueError("No data rows found in PDF")
        
        logger.info(f"Extracted {len(rows)} rows from PDF")
        return meta, rows
        
    except Exception as e:
        logger.error(f"Error extracting data from PDF: {str(e)}")
        raise

@app.route('/')
def index():
    """Main upload page."""
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert_pdf():
    """Handle PDF conversion."""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        if not allowed_file(file.filename):
            flash('Please upload a PDF file', 'error')
            return redirect(url_for('index'))
        
        # Secure the filename
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save uploaded file
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        logger.info(f"Saved uploaded file to: {pdf_path}")
        
        # Extract data from PDF
        meta, rows = extract_data_from_pdf(pdf_path)
        
        # Create output Excel file
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        # Create Excel file
        create_excel_from_rows(rows, meta, excel_path)
        
        logger.info(f"Created Excel file: {excel_path}")
        
        # Clean up uploaded PDF
        try:
            os.remove(pdf_path)
        except OSError:
            logger.warning(f"Could not remove uploaded file: {pdf_path}")
        
        # Return the Excel file
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"converted_{filename.rsplit('.', 1)[0]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error during conversion: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/api/convert', methods=['POST'])
def api_convert():
    """API endpoint for programmatic access."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        # Process file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        # Extract and convert
        meta, rows = extract_data_from_pdf(pdf_path)
        
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        create_excel_from_rows(rows, meta, excel_path)
        
        # Clean up
        try:
            os.remove(pdf_path)
        except OSError:
            pass
        
        # Return success response
        return jsonify({
            'success': True,
            'message': 'File converted successfully',
            'metadata': meta,
            'rows_processed': len(rows) - 1,  # Subtract header row
            'download_url': f'/download/{excel_filename}'
        })
        
    except Exception as e:
        logger.error(f"API conversion error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """Download converted files."""
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'error': 'Error downloading file'}), 500

@app.route('/health')
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)rom flask import Flask, request, jsonify, send_file, render_template_string, flash, redirect, url_for
from werkzeug.utils import secure_filename
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
import pdfplumber

# Setup basic logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'pdf-converter-secret-key-2025'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

ALLOWED_EXTENSIONS = {'pdf'}
UPLOAD_FOLDER = 'temp_uploads'
OUTPUT_FOLDER = 'temp_outputs'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# HTML template embedded in code to avoid file issues
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PDF to Excel Converter</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
        .container { background: white; border-radius: 20px; box-shadow: 0 20px 40px rgba(0, 0, 0, 0.1); padding: 40px; max-width: 600px; width: 100%; text-align: center; }
        .header h1 { color: #333; font-size: 2.5rem; margin-bottom: 10px; background: linear-gradient(135deg, #667eea, #764ba2); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
        .header p { color: #666; font-size: 1.1rem; margin-bottom: 30px; }
        .upload-area { border: 3px dashed #667eea; border-radius: 15px; padding: 40px 20px; margin: 30px 0; transition: all 0.3s ease; }
        .upload-area:hover { border-color: #764ba2; background: rgba(102, 126, 234, 0.05); }
        .upload-icon { font-size: 3rem; margin-bottom: 15px; }
        .upload-text { color: #333; font-size: 1.2rem; margin-bottom: 15px; }
        .upload-subtext { color: #666; font-size: 0.9rem; margin-bottom: 20px; }
        .file-input { display: none; }
        .browse-btn { background: linear-gradient(135deg, #667eea, #764ba2); color: white; border: none; padding: 12px 30px; border-radius: 25px; font-size: 1rem; cursor: pointer; transition: all 0.3s ease; margin: 10px 5px; }
        .browse-btn:hover { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4); }
        .convert-btn { background: linear-gradient(135deg, #56ab2f, #a8e6cf); color: white; border: none; padding: 15px 40px; border-radius: 25px; font-size: 1.1rem; cursor: pointer; transition: all 0.3s ease; margin-top: 20px; display: none; }
        .convert-btn:hover:not(:disabled) { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(86, 171, 47, 0.4); }
        .file-info { background: #f8f9fa; border-radius: 10px; padding: 15px; margin: 20px 0; display: none; }
        .file-info.show { display: block; }
        .flash-messages { margin-bottom: 20px; }
        .flash-message { padding: 12px 20px; border-radius: 8px; margin-bottom: 10px; font-weight: 500; }
        .flash-error { background: #fee; border: 1px solid #fcc; color: #c33; }
        .flash-success { background: #efe; border: 1px solid #cfc; color: #3c3; }
        .loading { display: none; align-items: center; justify-content: center; margin: 20px 0; }
        .loading.show { display: flex; }
        .spinner { width: 40px; height: 40px; border: 4px solid #e0e0e0; border-top: 4px solid #667eea; border-radius: 50%; animation: spin 1s linear infinite; margin-right: 15px; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>PDF to Excel Converter</h1>
            <p>Convert your PDF order documents to Excel spreadsheets</p>
        </div>

        <div class="flash-messages">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="flash-message flash-{{ category }}">{{ message }}</div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
        </div>

        <form id="uploadForm" action="/convert" method="post" enctype="multipart/form-data">
            <div class="upload-area" id="uploadArea">
                <div class="upload-icon">📄</div>
                <div class="upload-text">Drop your PDF file here</div>
                <div class="upload-subtext">or click to browse (Max 16MB)</div>
                <input type="file" id="fileInput" name="file" class="file-input" accept=".pdf" required>
                <button type="button" class="browse-btn" onclick="document.getElementById('fileInput').click()">Browse Files</button>
            </div>

            <div class="file-info" id="fileInfo">
                <strong>Selected file:</strong> <span id="fileName"></span><br>
                <strong>Size:</strong> <span id="fileSize"></span>
            </div>

            <div class="loading" id="loading">
                <div class="spinner"></div>
                <span>Processing your PDF...</span>
            </div>

            <button type="submit" class="convert-btn" id="convertBtn">Convert to Excel</button>
        </form>
    </div>

    <script>
        const fileInput = document.getElementById('fileInput');
        const fileInfo = document.getElementById('fileInfo');
        const fileName = document.getElementById('fileName');
        const fileSize = document.getElementById('fileSize');
        const convertBtn = document.getElementById('convertBtn');
        const uploadForm = document.getElementById('uploadForm');
        const loading = document.getElementById('loading');

        fileInput.addEventListener('change', function() {
            const file = fileInput.files[0];
            if (file) {
                if (file.type !== 'application/pdf') {
                    alert('Please select a PDF file.');
                    return;
                }
                if (file.size > 16 * 1024 * 1024) {
                    alert('File too large. Maximum size is 16MB.');
                    return;
                }
                fileName.textContent = file.name;
                fileSize.textContent = formatFileSize(file.size);
                fileInfo.classList.add('show');
                convertBtn.style.display = 'inline-block';
            }
        });

        uploadForm.addEventListener('submit', function() {
            convertBtn.disabled = true;
            convertBtn.textContent = 'Converting...';
            loading.classList.add('show');
        });

        function formatFileSize(bytes) {
            if (bytes === 0) return '0 Bytes';
            const k = 1024;
            const sizes = ['Bytes', 'KB', 'MB', 'GB'];
            const i = Math.floor(Math.log(bytes) / Math.log(k));
            return parseFloat((bytes / Math.pow(k, i)).toFixed(2)) + ' ' + sizes[i];
        }
    </script>
</body>
</html>
'''

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_header_meta(pdf_path):
    """Extract metadata from PDF."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return {}
            first_page = pdf.pages[0]
            text = first_page.extract_text() or ""

            po_match = re.search(r'PO ?(\d{6,})', text)
            ship_match = re.search(r'SHIP COMPLETE BY DATE:\s*(\d{1,2}/\d{1,2}/\d{4})', text)
            terms_match = re.search(r'PAYMENT TERMS:\s*(.*)', text)
            
            vendor_number = "N/A"
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if "Vendor #" in line:
                    vendor_match = re.search(r'Vendor #\s*(\d+)', line)
                    if vendor_match:
                        vendor_number = vendor_match.group(1)
                        break
                    elif i + 1 < len(lines):
                        next_line = lines[i+1].strip()
                        vendor_match = re.match(r'^(\d+)', next_line)
                        if vendor_match:
                            vendor_number = vendor_match.group(1)
                            break

            total_match = None
            for page in reversed(pdf.pages):
                page_text = page.extract_text() or ""
                total_match = re.search(r'Total\s+\$(\d{1,3}(?:,\d{3})*\.\d{2})', page_text)
                if total_match:
                    break

            return {
                "po_number": po_match.group(1) if po_match else "N/A",
                "vendor_number": vendor_number,
                "ship_by_date": ship_match.group(1) if ship_match else "N/A",
                "payment_terms": terms_match.group(1).strip() if terms_match else "N/A",
                "total": total_match.group(1) if total_match else "N/A",
                "page_count": len(pdf.pages)
            }
    except Exception as e:
        logger.error(f"Error extracting metadata: {e}")
        return {}

def extract_table_rows(pdf_path):
    """Extract table rows from PDF."""
    try:
        header_columns = ["Qty", "Item_SKU", "Dev_Code", "UPC", "HTS_Code", "Brand", "Description", "Rate", "Amount"]
        all_content_lines = []
        
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if not text:
                    continue
                
                lines = text.split('\n')
                for line in lines:
                    clean_line = line.strip()
                    if not clean_line or "Qty" in clean_line and "Item SKU" in clean_line or "This Purchase Order" in clean_line or "Page " in clean_line or clean_line.startswith("Total") or "TERMS:" in clean_line:
                        continue
                    all_content_lines.append(clean_line)

        full_text = " ".join(all_content_lines)
        
        # Regex patterns for different row formats
        patterns = [
            re.compile(r"(\d+)\s+([A-Z0-9-]+)\s+([A-Z0-9]+)\s+(\d{12})\s+(\d{10})\s+([A-Za-z][A-Za-z0-9\s&'-]+?)\s+(\$\d{1,3}(?:,\d{3})*\.\d{2})\s+(\$\d{1,3}(?:,\d{3})*\.\d{2})"),
            re.compile(r"(\d+)\s+([A-Z0-9-]+)\s+([A-Z0-9]+)\s+(\d{12})\s+([A-Za-z][A-Za-z0-9\s&'-]+?)\s+(\$\d{1,3}(?:,\d{3})*\.\d{2})\s+(\$\d{1,3}(?:,\d{3})*\.\d{2})")
        ]
        
        final_rows = [header_columns]
        found_items = set()
        
        for pattern_idx, pattern in enumerate(patterns):
            matches = pattern.findall(full_text)
            
            for match in matches:
                try:
                    if pattern_idx == 0:  # With HTS code
                        qty, sku, dev_code, upc, hts_code, brand_desc, rate, amount = match
                    else:  # Without HTS code
                        qty, sku, dev_code, upc, brand_desc, rate, amount = match
                        hts_code = ""
                    
                    brand_desc_parts = brand_desc.strip().split(None, 1)
                    brand = brand_desc_parts[0] if brand_desc_parts else ""
                    description = brand_desc_parts[1] if len(brand_desc_parts) > 1 else ""
                    
                    identifier = f"{qty}-{sku}-{dev_code}-{upc}"
                    if identifier in found_items:
                        continue
                    found_items.add(identifier)
                    
                    row = [
                        qty, sku, dev_code, upc, hts_code,
                        brand, description,
                        rate.replace('$', '').replace(',', ''),
                        amount.replace('$', '').replace(',', '')
                    ]
                    
                    if len(row) == len(header_columns):
                        final_rows.append(row)
                        
                except Exception as e:
                    logger.warning(f"Error processing match: {e}")
                    continue

        return final_rows
    except Exception as e:
        logger.error(f"Error extracting table rows: {e}")
        return []

def create_excel_from_rows(rows, meta, output_path):
    """Create Excel file from rows and metadata."""
    wb = Workbook()
    
    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    summary_data = [
        ['Field', 'Value'],
        ['PO Number', meta.get('po_number', 'N/A')],
        ['Vendor Number', meta.get('vendor_number', 'N/A')],
        ['Ship By Date', meta.get('ship_by_date', 'N/A')],
        ['Payment Terms', meta.get('payment_terms', 'N/A')],
        ['Total Amount', f"${meta.get('total', 'N/A')}" if meta.get('total') != 'N/A' else 'N/A'],
        ['Page Count', str(meta.get('page_count', 'N/A'))],
        ['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Format summary
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 30
    
    # Orders sheet
    orders_ws = wb.create_sheet(title="Orders")
    
    for row in rows:
        orders_ws.append(row)
    
    if rows:
        for cell in orders_ws[1]:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in orders_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        if column_letter == 'G':  # Description
            adjusted_width = min(max(max_length + 2, 20), 50)
        else:
            adjusted_width = max(max_length + 2, 12)
            
        orders_ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)

def extract_data_from_pdf(pdf_path):
    """Main extraction function."""
    meta = extract_header_meta(pdf_path)
    rows = extract_table_rows(pdf_path)
    
    if not rows or len(rows) <= 1:
        raise ValueError("No data rows found in PDF")
    
    return meta, rows

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/convert', methods=['POST'])
def convert_pdf():
    try:
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        if not allowed_file(file.filename):
            flash('Please upload a PDF file', 'error')
            return redirect(url_for('index'))
        
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        meta, rows = extract_data_from_pdf(pdf_path)
        
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        create_excel_from_rows(rows, meta, excel_path)
        
        try:
            os.remove(pdf_path)
        except OSError:
            pass
        
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"converted_{filename.rsplit('.', 1)[0]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Conversion error: {e}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/api/convert', methods=['POST'])
def api_convert():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(pdf_path)
        
        meta, rows = extract_data_from_pdf(pdf_path)
        
        excel_filename = f"{timestamp}_converted_{filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        create_excel_from_rows(rows, meta, excel_path)
        
        try:
            os.remove(pdf_path)
        except OSError:
            pass
        
        return jsonify({
            'success': True,
            'message': 'File converted successfully',
            'metadata': meta,
            'rows_processed': len(rows) - 1,
            'download_url': f'/download/{excel_filename}'
        })
        
    except Exception as e:
        logger.error(f"API conversion error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Download error: {e}")
        return jsonify({'error': 'Error downloading file'}), 500

@app.route('/health')
def health_check():
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
